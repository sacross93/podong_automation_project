from tkinter import Tk, Frame, Listbox, Entry, Button, Text, END, Label, messagebox, simpledialog, Scrollbar
from openpyxl import Workbook, load_workbook
import pyperclip

class CategoryWindow:
    def __init__(self, master):
        self.master = master
        self.master.title("Category Window")

        self.category_listbox = Listbox(self.master)
        self.category_listbox.pack(side="left", padx=10, pady=10, fill="both", expand=True)

        scrollbar = Scrollbar(self.master)
        scrollbar.pack(side="right", padx=(0, 10), pady=10, fill="y")

        self.category_listbox.config(yscrollcommand=scrollbar.set)
        scrollbar.config(command=self.category_listbox.yview)

        frame_buttons = Frame(self.master)
        frame_buttons.pack(padx=10, pady=(0, 10))

        self.select_button = Button(frame_buttons, text="Select", command=self.select_category)
        self.select_button.pack(side="left", padx=5)

        self.delete_button = Button(frame_buttons, text="Delete", command=self.delete_category)
        self.delete_button.pack(side="left", padx=5)

        self.rename_button = Button(frame_buttons, text="Rename", command=self.rename_category)
        self.rename_button.pack(side="left", padx=5)

        self.selected_category = None

    def update_categories(self, categories):
        self.category_listbox.delete(0, END)
        for category in categories:
            self.category_listbox.insert(END, category)
        self.category_listbox.yview_moveto(0)  # Set scrollbar position to top

    def select_category(self):
        selected_index = self.category_listbox.curselection()
        if selected_index:
            self.selected_category = self.category_listbox.get(selected_index)
            content_window.update_content(data.get(self.selected_category, ""))
            status_label.config(text=f"Selected Category: {self.selected_category}")
        else:
            status_label.config(text="No Category Selected")

    def delete_category(self):
        selected_index = self.category_listbox.curselection()
        if selected_index:
            selected_category = self.category_listbox.get(selected_index)
            response = messagebox.askyesno(
                "Delete Category", f"Are you sure you want to delete the category '{selected_category}'?"
            )
            if response:
                del data[selected_category]
                save_data_to_excel()
                self.update_categories(data.keys())
                content_window.update_content("")
                status_label.config(text="Category Deleted")
        else:
            status_label.config(text="No Category Selected")

    def rename_category(self):
        selected_index = self.category_listbox.curselection()
        if selected_index:
            selected_category = self.category_listbox.get(selected_index)
            new_category = simpledialog.askstring(
                "Rename Category", f"Enter a new name for the category '{selected_category}':"
            )
            if new_category:
                data[new_category] = data.pop(selected_category)
                save_data_to_excel()
                self.update_categories(data.keys())
                content_window.update_content(data.get(new_category, ""))
                status_label.config(text=f"Category '{selected_category}' renamed to '{new_category}'")
        else:
            status_label.config(text="No Category Selected")

class AddCategoryWindow:
    def __init__(self, master, category_window):
        self.master = master
        self.category_window = category_window
        self.master.title("Add Category Window")

        self.category_entry = Entry(self.master)
        self.category_entry.pack(padx=10, pady=10)

        self.add_button = Button(self.master, text="Add", command=self.add_category)
        self.add_button.pack(padx=10, pady=(0, 10))

    def add_category(self):
        category = self.category_entry.get().strip()
        if category:
            data[category] = ""
            save_data_to_excel()
            self.category_window.update_categories(data.keys())
            self.category_entry.delete(0, END)
            status_label.config(text="Category Added")
        else:
            status_label.config(text="Empty Category")

class ContentWindow:
    def __init__(self, master):
        self.master = master
        self.master.title("Content Window")

        self.content_text = Text(self.master, height=10)
        self.content_text.pack(padx=10, pady=10)

        frame_buttons = Frame(self.master)
        frame_buttons.pack(padx=10, pady=(0, 10))

        self.save_button = Button(frame_buttons, text="Save", command=self.save_content)
        self.save_button.pack(side="left", padx=5)

        self.copy_button = Button(frame_buttons, text="Copy", command=self.copy_content)
        self.copy_button.pack(side="left", padx=5)

    def update_content(self, content):
        self.content_text.delete(1.0, END)
        self.content_text.insert(END, content)

    def save_content(self):
        category = category_window.selected_category
        if category:
            content = self.content_text.get(1.0, END).strip()
            if content != data.get(category, ""):
                data[category] = content
                save_data_to_excel()
                status_label.config(text="Content Saved")
            else:
                status_label.config(text="No changes made")
        else:
            status_label.config(text="No Category Selected")

    def copy_content(self):
        category = category_window.selected_category
        if category:
            content = data.get(category, "")
            pyperclip.copy(content)
            status_label.config(text="Content Copied to Clipboard")
        else:
            status_label.config(text="No Category Selected")


def load_data_from_excel():
    try:
        workbook = load_workbook("data.xlsx")
        sheet = workbook.active

        for row in sheet.iter_rows(min_row=2, values_only=True):
            category = row[0]
            content = row[1]
            data[category] = content

        category_window.update_categories(data.keys())
        status_label.config(text="Data Loaded")
    except FileNotFoundError:
        status_label.config(text="Data File Not Found")
    except Exception as e:
        status_label.config(text=f"Error Loading Data: {str(e)}")


def save_data_to_excel():
    try:
        workbook = Workbook()
        sheet = workbook.active

        sheet.append(["Category", "Content"])

        for category, content in data.items():
            sheet.append([category, content])

        workbook.save("data.xlsx")
        status_label.config(text="Data Saved")
    except Exception as e:
        status_label.config(text=f"Error Saving Data: {str(e)}")


if __name__ == "__main__":
    root = Tk()
    root.geometry("400x520")
    root.configure(bg="white")

    data = {}  # Dictionary to store categories and their corresponding content

    category_window = CategoryWindow(root)
    add_category_window = AddCategoryWindow(root, category_window)
    content_window = ContentWindow(root)

    status_label = Label(root, text="", bg="white")
    status_label.pack(padx=10, pady=(0, 10))

    load_data_from_excel()

    root.mainloop()