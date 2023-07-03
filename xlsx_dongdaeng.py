import tkinter as tk
from openpyxl import load_workbook, Workbook

def load_excel_data(file_path):
    workbook = load_workbook(file_path)
    sheet = workbook.active
    data = []

    for row in sheet.iter_rows(min_row=2, values_only=True):
        data.append(row)

    workbook.close()  # Close the workbook after reading the data
    return data

def create_title_buttons(data):
    def on_button_click(title):
        content_text.delete('1.0', tk.END)  # Clear the content text area
        for row in data:
            if row[0] == title:
                content_text.insert(tk.END, row[1])  # Display the content

    for row in data:
        title = row[0]
        button = tk.Button(root, text=title, command=lambda title=title: on_button_click(title))
        button.pack(side=tk.TOP)

def add_title_and_content():
    title = title_entry.get()
    content = content_text.get('1.0', tk.END)

    # Save the title and content to Excel file
    workbook = load_workbook(file_path)
    sheet = workbook.active
    sheet.append([title, content])
    workbook.save(file_path)
    workbook.close()  # Close the workbook after saving the data

    # Create a new button for the added title
    button = tk.Button(root, text=title, command=lambda title=title: on_button_click(title))
    button.pack(side=tk.TOP)

# Create the main window
root = tk.Tk()
root.title("Excel Automation Program")

# Load Excel data
file_path = "./file.xlsx"
data = load_excel_data(file_path)

# Create title buttons
create_title_buttons(data)

# Add title and content entry fields
title_label = tk.Label(root, text="Title:")
title_label.pack(side=tk.TOP)
title_entry = tk.Entry(root)
title_entry.pack(side=tk.TOP)

content_label = tk.Label(root, text="Content:")
content_label.pack(side=tk.TOP)
content_text = tk.Text(root)
content_text.pack(side=tk.TOP)

# Add title and content button
add_button = tk.Button(root, text="Add Title and Content", command=add_title_and_content)
add_button.pack(side=tk.TOP)

# Start the GUI event loop
root.mainloop()